#!/usr/bin/env python3
"""Dump inputs, formulas, and values from Free Model Excel workbooks."""
import json
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parents[1] / "FINMECH-UPGRADED" / "1.Free Models"

FILES = [
    "Revenue Model.xlsx",
    "Costing Model.xlsx",
    "Break-even Model- Only calculator.xlsx",
    "Know your Business Numbers.xlsx",
]


def cell_info(ws, row, col):
    c = ws.cell(row=row, column=col)
    return {
        "coord": c.coordinate,
        "value": c.value,
        "formula": c.value if isinstance(c.value, str) and c.value.startswith("=") else None,
        "display": c.value,
    }


def dump_sheet(ws, max_row=80, max_col=12):
    rows = []
    for r in range(1, min(ws.max_row or 1, max_row) + 1):
        row_cells = []
        for c in range(1, min(ws.max_column or 1, max_col) + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                row_cells.append(
                    {
                        "col": c,
                        "coord": cell.coordinate,
                        "value": cell.value,
                    }
                )
        if row_cells:
            rows.append({"row": r, "cells": row_cells})
    return rows


def main():
    out = {}
    for fname in FILES:
        path = BASE / fname
        if not path.exists():
            print(f"MISSING: {path}", file=sys.stderr)
            continue
        wb = openpyxl.load_workbook(path, data_only=False)
        wb_vals = openpyxl.load_workbook(path, data_only=True)
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            ws_v = wb_vals[name]
            # merged values view
            data_rows = []
            for r in range(1, min(ws.max_row or 1, 100) + 1):
                for c in range(1, min(ws.max_column or 1, 15) + 1):
                    raw = ws.cell(row=r, column=c).value
                    val = ws_v.cell(row=r, column=c).value
                    if raw is not None or val is not None:
                        data_rows.append(
                            {
                                "coord": ws.cell(row=r, column=c).coordinate,
                                "label_or_formula": raw,
                                "computed_value": val,
                            }
                        )
            sheets[name] = data_rows
        out[fname] = {"sheets": list(wb.sheetnames), "cells": sheets}

    print(json.dumps(out, indent=2, default=str))


if __name__ == "__main__":
    main()
