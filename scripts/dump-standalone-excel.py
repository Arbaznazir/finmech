#!/usr/bin/env python3
"""Dump standalone Excel workbooks for verification."""
import json
import sys
from pathlib import Path
import openpyxl

BASE = Path(__file__).resolve().parents[1] / "FINMECH-UPGRADED" / "2.Stand alone models"
FILES = [
    "1.Income Statement.xlsx",
    "2.Balance Sheet.xlsx",
    "3.Burn and Runway Monitor.xlsx",
]


def dump_workbook(path):
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)
    out = {"sheets": wb_f.sheetnames, "data": {}}
    for name in wb_f.sheetnames:
        ws, ws_v = wb_f[name], wb_v[name]
        rows = []
        for r in range(1, min(ws.max_row or 1, 120) + 1):
            row = []
            for c in range(1, min(ws.max_column or 1, 20) + 1):
                raw = ws.cell(r, c).value
                val = ws_v.cell(r, c).value
                if raw is not None or val is not None:
                    row.append({
                        "coord": ws.cell(r, c).coordinate,
                        "raw": raw,
                        "value": val,
                    })
            if row:
                rows.append({"row": r, "cells": row})
        out["data"][name] = rows
    return out


def main():
    result = {}
    for f in FILES:
        p = BASE / f
        if p.exists():
            result[f] = dump_workbook(p)
        else:
            result[f] = {"error": "missing"}
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()
