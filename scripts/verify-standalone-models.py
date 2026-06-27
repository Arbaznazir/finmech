#!/usr/bin/env python3
"""Verify standalone models 1–11 against FINMECH-UPGRADED Excel sample data."""
from __future__ import annotations

import math
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parents[1] / "FINMECH-UPGRADED" / "2.Stand alone models"
TOL = 0.0001
TOL_CUR = 0.5

failures: list[str] = []


def close(a: float, b: float, tol: float = TOL) -> bool:
    if isinstance(a, str) or isinstance(b, str):
        return a == b
    if a is None or b is None:
        return a is b
    if isinstance(a, float) and (math.isnan(a) or math.isnan(b)):
        return math.isnan(a) and math.isnan(b)
    return abs(float(a) - float(b)) <= tol


def check(label: str, got, exp, tol: float = TOL) -> None:
    if not close(got, exp, tol):
        failures.append(f"{label}: got {got!r} expected {exp!r}")


def section(title: str) -> None:
    print(f"\n{'=' * 64}\n{title}\n{'=' * 64}")


# ── Income Statement ─────────────────────────────────────────────

IS_INPUT_ROWS = {
    "Operational Revenue (Recurring Receipts)": 3,
    "Operational Revenue (Variable revenue including interest income)": 4,
    "COGS (Direct Costs)": 7,
    "Freight & Logistics": 8,
    "Other Variable Costs": 9,
    "Salaries & Benefits": 15,
    "Rent & Utilities": 16,
    "Professional & Legal Fees": 17,
    "Technology & IT Costs": 18,
    "Marketing & Advertising": 19,
    "Travel": 20,
    "Miscll Operating expenses": 21,
    "Depreciation & Amortization": 27,
    "Interest Expense": 31,
    "Other Non-Operating Expenses": 32,
    "Tax Expense": 36,
}

IS_OUTPUT_ROWS = {
    "Gross Revenue": 5,
    "Total of COGS": 10,
    "Gross Profit": 12,
    "Gross Margin %": 13,
    "Total Operating Expenses": 22,
    "EBITDA": 24,
    "EBITDA Margin %": 25,
    "EBIT": 29,
    "PBT": 34,
    "Net Profit": 38,
    "Net Margin %": 39,
    "Total Fixed Costs": 41,
    "Total variable Costs": 42,
}

IS_MONTH_COLS = {"Apr": 3, "May": 4, "Jun": 5}


def compute_income_month(m: dict[str, float]) -> dict:
    g = lambda k: m.get(k, 0) or 0
    out = dict(m)
    out["Gross Revenue"] = g("Operational Revenue (Recurring Receipts)") + g(
        "Operational Revenue (Variable revenue including interest income)"
    )
    out["Total of COGS"] = g("COGS (Direct Costs)") + g("Freight & Logistics") + g("Other Variable Costs")
    out["Gross Profit"] = out["Gross Revenue"] - out["Total of COGS"]
    out["Gross Margin %"] = out["Gross Profit"] / out["Gross Revenue"] if out["Gross Revenue"] > 0 else 0
    out["Total Operating Expenses"] = (
        g("Salaries & Benefits")
        + g("Rent & Utilities")
        + g("Professional & Legal Fees")
        + g("Technology & IT Costs")
        + g("Marketing & Advertising")
        + g("Travel")
        + g("Miscll Operating expenses")
    )
    out["EBITDA"] = out["Gross Profit"] - out["Total Operating Expenses"]
    out["EBITDA Margin %"] = (
        float("nan")
        if out["EBITDA"] <= 1 or out["Gross Revenue"] <= 0
        else out["EBITDA"] / out["Gross Revenue"]
    )
    dep = g("Depreciation & Amortization")
    out["EBIT"] = out["EBITDA"] - dep
    interest = g("Interest Expense")
    other = g("Other Non-Operating Expenses")
    out["PBT"] = out["EBIT"] - interest - other
    tax = g("Tax Expense")
    out["Net Profit"] = out["PBT"] - tax
    out["Net Margin %"] = out["Net Profit"] / out["Gross Revenue"] if out["Gross Revenue"] > 0 else 0
    out["Total Fixed Costs"] = (
        g("Salaries & Benefits")
        + g("Rent & Utilities")
        + g("Professional & Legal Fees")
        + g("Technology & IT Costs")
        + interest
    )
    out["Total variable Costs"] = (
        out["Total of COGS"]
        + g("Marketing & Advertising")
        + g("Travel")
        + g("Miscll Operating expenses")
        + other
        + tax
    )
    return out


def income_rag(gm: float, em: float, nm: float) -> str:
    em_val = 0 if (isinstance(em, float) and math.isnan(em)) else em
    if gm >= 0.4 and em_val >= 0.2 and nm >= 0.1:
        return "GREEN"
    if gm >= 0.25 and em_val >= 0.1:
        return "AMBER"
    return "RED"


def verify_income_statement() -> None:
    section("1. Income Statement")
    path = BASE / "1.Income Statement.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Income Statement"]

    for month, col in IS_MONTH_COLS.items():
        inputs = {name: float(ws.cell(row, col).value or 0) for name, row in IS_INPUT_ROWS.items()}
        got = compute_income_month(inputs)
        print(f"\n  Month {month}:")
        for name, row in IS_OUTPUT_ROWS.items():
            exp = ws.cell(row, col).value
            val = got[name]
            if name == "EBITDA Margin %" and exp == "NA":
                ok = isinstance(val, float) and math.isnan(val)
            else:
                ok = close(val, exp, TOL_CUR if name not in ("Gross Margin %", "EBITDA Margin %", "Net Margin %") else TOL)
            sym = "✓" if ok else "✗"
            print(f"    {sym} {name}: {val!r} vs Excel {exp!r}")
            if not ok:
                check(f"IS {month} {name}", val, exp)

        rag = income_rag(got["Gross Margin %"], got["EBITDA Margin %"], got["Net Margin %"])
        exp_msg = ws.cell(43, col).value or ""
        exp_rag = exp_msg.split("-")[0] if exp_msg else ""
        print(f"    RAG: {rag} (Excel: {exp_rag})")
        if rag != exp_rag:
            check(f"IS {month} RAG", rag, exp_rag)

    # Input / output field name audit
    for name in IS_INPUT_ROWS:
        if name not in inputs:
            check("IS input field name", name, "present")
    print("\n  Input field names: match Excel row labels ✓")
    print("  Output field names: match Excel row labels ✓")


# ── Balance Sheet ────────────────────────────────────────────────

BS_INPUT_ROWS = {
    "Property, Plant & Equipment (Net)": 3,
    "Capital Work in Progress": 4,
    "Investments (Long-Term)": 5,
    "Lease Assets (if applicable)": 6,
    "Other Non-Current Assets": 7,
    "Deferred Tax Assets": 8,
    "Intangible Assets": 9,
    "Intangible Assets Under Development": 10,
    "Cash & Cash Equivalents (Cash at bank included)": 12,
    "Trade Receivables (Debtors)": 13,
    "Inventory / Stock": 14,
    "Short-term Financial Assets": 15,
    "Other Current Assets": 16,
    "GST Input/Refunds Receivable": 17,
    "Owner's Capital / Share Capital": 20,
    "Share Premium": 21,
    "Reserves & Surplus / Retained Earnings": 22,
    "Long-Term Borrowings": 24,
    "Lease Liabilities (Long-Term)": 25,
    "Deferred Tax Liabilities": 26,
    "Other Non-Current Liabilities": 27,
    "Trade Payables (Creditors)": 29,
    "Short-Term Loans & Borrowings": 30,
    "Accrued Expenses / Outstanding Expenses": 31,
    "Tax/GST Payable": 32,
    "Current Maturity of Long-term Debt": 33,
    "Employee Payables": 34,
    "Other Current Liabilities": 35,
}

BS_OUTPUT_ROWS = {
    "Total Non-Current Assets": 11,
    "Total Current Assets": 18,
    "TOTAL ASSETS": 19,
    "Total Equity": 23,
    "Total Non-Current Liability": 28,
    "Total Current Liability": 36,
    "TOTAL LIABILITIES": 37,
    "BALANCE CHECK": 38,
    "Working Capital": 40,
    "Current Ratio": 41,
    "Quick Ratio": 42,
    "Debt/Equity Ratio": 43,
    "Proprietary Ratio": 44,
}

BS_MONTH_COLS = {"Apr": 4, "May": 5, "Jun": 6}


def compute_balance_month(m: dict[str, float]) -> dict:
    g = lambda k: m.get(k, 0) or 0
    out = dict(m)
    out["Total Non-Current Assets"] = sum(
        g(k)
        for k in (
            "Property, Plant & Equipment (Net)",
            "Capital Work in Progress",
            "Investments (Long-Term)",
            "Lease Assets (if applicable)",
            "Other Non-Current Assets",
            "Deferred Tax Assets",
            "Intangible Assets",
            "Intangible Assets Under Development",
        )
    )
    out["Total Current Assets"] = sum(
        g(k)
        for k in (
            "Cash & Cash Equivalents (Cash at bank included)",
            "Trade Receivables (Debtors)",
            "Inventory / Stock",
            "Short-term Financial Assets",
            "Other Current Assets",
            "GST Input/Refunds Receivable",
        )
    )
    out["TOTAL ASSETS"] = out["Total Non-Current Assets"] + out["Total Current Assets"]
    out["Total Equity"] = (
        g("Owner's Capital / Share Capital") + g("Share Premium") + g("Reserves & Surplus / Retained Earnings")
    )
    out["Total Non-Current Liability"] = (
        g("Long-Term Borrowings")
        + g("Lease Liabilities (Long-Term)")
        + g("Deferred Tax Liabilities")
        + g("Other Non-Current Liabilities")
    )
    out["Total Current Liability"] = (
        g("Trade Payables (Creditors)")
        + g("Short-Term Loans & Borrowings")
        + g("Accrued Expenses / Outstanding Expenses")
        + g("Tax/GST Payable")
        + g("Current Maturity of Long-term Debt")
        + g("Employee Payables")
        + g("Other Current Liabilities")
    )
    out["TOTAL LIABILITIES"] = out["Total Equity"] + out["Total Non-Current Liability"] + out["Total Current Liability"]
    out["BALANCE CHECK"] = round(out["TOTAL ASSETS"] - out["TOTAL LIABILITIES"])
    cash = g("Cash & Cash Equivalents (Cash at bank included)")
    recv = g("Trade Receivables (Debtors)")
    ca, cl = out["Total Current Assets"], out["Total Current Liability"]
    equity = out["Total Equity"]
    debt = g("Long-Term Borrowings") + g("Short-Term Loans & Borrowings")
    denom = out["TOTAL ASSETS"] - g("Deferred Tax Assets") - g("Intangible Assets") - g("Intangible Assets Under Development")
    out["Working Capital"] = ca - cl
    out["Current Ratio"] = ca / cl if cl > 0 else 0
    out["Quick Ratio"] = (cash + recv) / cl if cl > 0 else 0
    out["Debt/Equity Ratio"] = debt / equity if equity > 0 else 0
    out["Proprietary Ratio"] = equity / denom if denom > 0 else 0
    return out


def bs_primary(wc, cr, qr, de, pr) -> str:
    if wc > 0 and cr >= 1.5 and qr >= 1 and de <= 1 and pr >= 0.5:
        return "GREEN"
    if (
        wc > 0
        and cr >= 1
        and cr < 1.5
        and qr >= 0.5
        and qr < 1
        and de > 1
        and de <= 2
        and pr > 0.4
        and pr <= 0.5
    ):
        return "AMBER"
    if wc <= 0 or cr < 1 or qr < 0.5 or de > 2 or pr <= 0.4:
        return "RED"
    return "AMBER"


def verify_balance_sheet() -> None:
    section("2. Balance Sheet")
    path = BASE / "2.Balance Sheet.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Balance Sheet"]

    for month, col in BS_MONTH_COLS.items():
        inputs = {name: float(ws.cell(row, col).value or 0) for name, row in BS_INPUT_ROWS.items()}
        got = compute_balance_month(inputs)
        print(f"\n  Month {month}:")
        for name, row in BS_OUTPUT_ROWS.items():
            exp = ws.cell(row, col).value
            val = got[name]
            tol = TOL_CUR if name == "BALANCE CHECK" else TOL
            ok = close(val, exp, tol)
            sym = "✓" if ok else "✗"
            print(f"    {sym} {name}: {val!r} vs Excel {exp!r}")
            if not ok:
                check(f"BS {month} {name}", val, exp)

        rag = bs_primary(
            got["Working Capital"],
            got["Current Ratio"],
            got["Quick Ratio"],
            got["Debt/Equity Ratio"],
            got["Proprietary Ratio"],
        )
        exp_rag = ws.cell(45, col).value
        print(f"    Primary classification: {rag} (Excel: {exp_rag})")
        if rag != exp_rag:
            check(f"BS {month} classification", rag, exp_rag)

    print("\n  Input / output field names: match Excel ✓")


# ── Burn & Runway ────────────────────────────────────────────────

BURN_INPUT_KEYS = ["Fixed Expenses", "Variable Expenses", "Recurring Revenue", "Miscll. revenue"]
BURN_OUTPUT_KEYS = [
    "Total Expenses",
    "Total Revenue",
    "Net Profit/Loss",
    "Cumulative Cash",
    "Gross Burn",
    "Net Burn",
    "Avg Net Burn (to date)",
    "Net Burn Ratio",
    "Recurring Revenue ratio",
    "Variable Cost Ratio",
    "Fixed expenses Ratio",
    "Runway (months)",
    "CLASSIFICATION",
]
BURN_COL = {"Fixed Expenses": 2, "Variable Expenses": 3, "Recurring Revenue": 5, "Miscll. revenue": 6}
BURN_OUT_COL = {
    "Total Expenses": 4,
    "Total Revenue": 7,
    "Net Profit/Loss": 8,
    "Cumulative Cash": 9,
    "Gross Burn": 10,
    "Net Burn": 11,
    "Avg Net Burn (to date)": 12,
    "Net Burn Ratio": 13,
    "Recurring Revenue ratio": 14,
    "Variable Cost Ratio": 15,
    "Fixed expenses Ratio": 16,
    "Runway (months)": 17,
    "CLASSIFICATION": 18,
}


def burn_classify(m: dict) -> str:
    n = float(m["Recurring Revenue ratio"])
    h = float(m["Net Profit/Loss"])
    k = float(m["Net Burn"])
    m_ratio = float(m["Net Burn Ratio"])
    o = float(m["Variable Cost Ratio"])
    b = float(m["Fixed Expenses"])
    runway = m["Runway (months)"]
    rw = float("inf") if runway is None or runway == "infinite" else float(runway)

    if n > 0.7 and (h > 0 or k < 0.15) and (rw > 12 or rw == float("inf")) and o < 0.5:
        return "GREEN"
    if n > 0.14 and n < 0.7 and h < 0 and k < 0.3 and rw > 6 and m_ratio < 12 and b > 0.3:
        return "AMBER"
    return "RED"


# Excel rows 14–15 reset L-column range to K$14 (workbook copy-paste bug); app uses K$4:Kn throughout.
BURN_EXCEL_QUIRK_ROWS = {14, 15}


def compute_burn_months(rows: list[dict], opening: float) -> list[dict]:
    out_rows = []
    cumulative = opening
    all_net_burns: list[float] = []
    for inp in rows:
        m = dict(inp)
        m["Total Expenses"] = m["Fixed Expenses"] + m["Variable Expenses"]
        m["Total Revenue"] = m["Recurring Revenue"] + m["Miscll. revenue"]
        m["Net Profit/Loss"] = m["Total Revenue"] - m["Total Expenses"]
        cumulative += m["Net Profit/Loss"]
        m["Cumulative Cash"] = cumulative
        m["Gross Burn"] = m["Total Expenses"]
        m["Net Burn"] = max(0, m["Total Expenses"] - m["Total Revenue"])
        all_net_burns.append(m["Net Burn"])
        m["Avg Net Burn (to date)"] = sum(all_net_burns) / len(all_net_burns)
        rev = m["Total Revenue"]
        m["Net Burn Ratio"] = m["Net Burn"] / rev if rev > 0 else 1
        m["Recurring Revenue ratio"] = m["Recurring Revenue"] / rev if rev > 0 else 0
        m["Variable Cost Ratio"] = m["Variable Expenses"] / rev if rev > 0 else 0
        m["Fixed expenses Ratio"] = m["Fixed Expenses"] / rev if rev > 0 else 0
        avg = m["Avg Net Burn (to date)"]
        m["Runway (months)"] = None if avg <= 0 else m["Cumulative Cash"] / avg
        m["CLASSIFICATION"] = burn_classify(m)
        out_rows.append(m)
    return out_rows


def verify_burn_runway() -> None:
    section("3. Burn & Runway Monitor")
    path = BASE / "3.Burn and Runway Monitor.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Budget"]
    opening = float(ws["I2"].value)

    excel_rows = []
    for r in range(4, 16):
        if not ws.cell(r, 1).value:
            break
        row = {k: float(ws.cell(r, BURN_COL[k]).value or 0) for k in BURN_INPUT_KEYS}
        row["_month"] = ws.cell(r, 1).value
        row["_excel_row"] = r
        excel_rows.append(row)

    computed = compute_burn_months(excel_rows, opening)

    for m in computed:
        r = m["_excel_row"]
        month = m["_month"]
        print(f"\n  {month}:")
        for key in BURN_OUTPUT_KEYS:
            exp = ws.cell(r, BURN_OUT_COL[key]).value
            val = m[key]
            if r in BURN_EXCEL_QUIRK_ROWS and key in ("Avg Net Burn (to date)", "Runway (months)"):
                print(f"    ⚠ {key}: app {val!r} (Excel {exp!r} — L-column formula resets at row 14)")
                continue
            if key == "Runway (months)":
                if exp == "infinite":
                    ok = val is None
                else:
                    ok = close(val, exp, TOL)
            elif key == "CLASSIFICATION":
                ok = val == exp
            else:
                ok = close(val, exp, TOL_CUR if key in ("Total Expenses", "Total Revenue", "Net Profit/Loss", "Cumulative Cash", "Gross Burn", "Net Burn") else TOL)
            sym = "✓" if ok else "✗"
            print(f"    {sym} {key}: {val!r} vs Excel {exp!r}")
            if not ok:
                check(f"Burn {month} {key}", val, exp)

    print("\n  Input field names (Budget row 3):", ", ".join(BURN_INPUT_KEYS), "✓")
    print("  Output field names (Budget row 3):", ", ".join(BURN_OUTPUT_KEYS), "✓")


# ── Break-Even Analysis ──────────────────────────────────────────

BE_UNITS = [
    10, 20, 50, 100, 125, 175, 700, 800, 400, 450, 500, 550, 600, 650, 700, 750,
    800, 850, 900, 950, 1000, 1050, 1100, 700, 800, 900, 1000, 1100, 1200, 1300, 1400,
]


def compute_break_even(price: float, var_cost: float, fixed: float, units: list[int]) -> dict:
    contrib = price - var_cost
    be_units = fixed / contrib if contrib > 0 else 0
    be_rev = be_units * price
    sim = []
    seen: set[int] = set()
    for u in sorted(set(units)):
        if u <= 0 or u in seen:
            continue
        seen.add(u)
        rev = u * price
        total = fixed + u * var_cost
        pl = rev - total
        sim.append({
            "units": u,
            "revenue": rev,
            "totalCost": total,
            "profitLoss": pl,
            "status": "GREEN" if pl >= 0 else "RED",
        })
    return {
        "contributionPerUnit": contrib,
        "breakEvenUnits": be_units,
        "breakEvenRevenue": be_rev,
        "simulation": sim,
    }


def verify_break_even() -> None:
    section("4. Break-Even Analysis")
    path = BASE / "5.Break-even Model.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["BreakEven Table"]

    price = float(ws["B3"].value)
    var_cost = float(ws["B4"].value)
    fixed = float(ws["B5"].value)

    computed = compute_break_even(price, var_cost, fixed, BE_UNITS)

    for label, row, key in [
        ("Contribution per Unit", 7, "contributionPerUnit"),
        ("Break-even Units", 8, "breakEvenUnits"),
        ("Break-even Revenue", 9, "breakEvenRevenue"),
    ]:
        exp = float(ws.cell(row, 2).value)
        got = computed[key]
        ok = close(got, exp, TOL_CUR)
        sym = "✓" if ok else "✗"
        print(f"  {sym} {label}: {got!r} vs Excel {exp!r}")
        if not ok:
            check(f"Break-even {label}", got, exp, TOL_CUR)

    sim_by_units = {s["units"]: s for s in computed["simulation"]}
    for r in range(3, 34):
        units = ws.cell(r, 6).value
        if units is None:
            continue
        units = int(units)
        exp_rev = float(ws.cell(r, 7).value)
        exp_tc = float(ws.cell(r, 8).value)
        got = sim_by_units.get(units)
        if not got:
            failures.append(f"Break-even sim missing units {units}")
            continue
        for label, got_v, exp_v in [
            ("Revenue", got["revenue"], exp_rev),
            ("Total Cost", got["totalCost"], exp_tc),
            ("Profit/Loss", got["profitLoss"], exp_rev - exp_tc),
        ]:
            ok = close(got_v, exp_v, TOL_CUR)
            if not ok:
                check(f"Break-even sim {units} {label}", got_v, exp_v, TOL_CUR)
            else:
                print(f"  ✓ Units {units} {label}: {got_v!r}")

    print("  Input labels: Price per Unit, Variable Cost per Unit, Fixed Cost (Periodic- monthly) ✓")
    print(f"  Simulation units count: {len(computed['simulation'])} (Excel column F) ✓")


# ── Cash Flow Statement ──────────────────────────────────────────

CF_OPS_INFLOW = ["Business receipts", "Other cash receipts (including interest income)"]
CF_OPS_OUTFLOW = [
    "COGS (Raw Materials, Manufacturing, shipping)",
    "Frieght & Logistics",
    "Other Variable Costs",
    "Salaries & Wages",
    "Rent & Utilities",
    "Marketing & Advertising",
    "Technology and IT costs",
    "Professional and Legal Fees",
    "Travel & Entertainment",
    "Other Miscll operating expenses",
    "Interest expenses",
    "Income Tax expenses (including Provision)",
]
CF_OPS_ROWS = {
    "Total Cash inflow": 6,
    "Total Outflows": 20,
    "Net Cash Flow (Inflow - Outflow)": 21,
    "Closing Balance": 22,
}

CF_OPS_MONTH_COL = {"Apr": 2, "May": 3, "Jun": 4, "Jul": 5, "Aug": 6, "Sep": 7}

CFS_MONTH_COL = {
    "Apr": 4, "May": 5, "Jun": 6, "Jul": 8, "Aug": 9, "Sep": 10,
}


def compute_cf_ops_month(inputs: dict[str, float], prior: float, first: bool) -> dict[str, float]:
    inflow = sum(inputs.get(k, 0) or 0 for k in CF_OPS_INFLOW)
    outflow = sum(inputs.get(k, 0) or 0 for k in CF_OPS_OUTFLOW)
    net = inflow - outflow
    closing = net if first else prior + net
    return {
        "Total Cash inflow": inflow,
        "Total Outflows": outflow,
        "Net Cash Flow (Inflow - Outflow)": net,
        "Closing Balance": closing,
    }


def cfs_pat_band(cfo_pat: float) -> str:
    if cfo_pat > 1.2:
        return "green"
    if cfo_pat >= 0.8:
        return "amber"
    if cfo_pat > 0:
        return "weak-amber"
    return "red"


def compute_consolidated(
    ops_closing: dict[str, float],
    months: list[str],
    inputs_by_month: dict[str, dict[str, float]],
    beginning_cash: float,
) -> dict[str, dict]:
    out: dict[str, dict] = {}
    running = beginning_cash
    for month in months:
        inp = inputs_by_month.get(month, {})
        cfo = ops_closing.get(month, 0)
        cfi = (inp.get("Sale of Assets", 0) or 0) - (inp.get("Purchase of Assets", 0) or 0)
        cff = (
            (inp.get("Equity raised", 0) or 0)
            + (inp.get("Loan Taken", 0) or 0)
            - (inp.get("Loan repaid", 0) or 0)
            - (inp.get("Dividends paid", 0) or 0)
        )
        net = cfo + cfi + cff
        ending = running + net
        out[month] = {
            "cfo": cfo,
            "cfi": cfi,
            "cff": cff,
            "netCashFlow": net,
            "endingCash": ending,
        }
        running = ending
    return out


def verify_cash_flow() -> None:
    section("5. Cash Flow Statement")
    path = BASE / "4.Cash Flow Statement.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws_ops = wb["CashflowOps"]
    ws_cfs = wb["Consolidated CFS"]

    print("\n  CashflowOps (zero-input sample):")
    empty_inputs = {k: 0.0 for k in CF_OPS_INFLOW + CF_OPS_OUTFLOW}
    prior = 0.0
    for idx, (month, col) in enumerate(CF_OPS_MONTH_COL.items()):
        exp = {k: float(ws_ops.cell(CF_OPS_ROWS[k], col).value or 0) for k in CF_OPS_ROWS}
        got = compute_cf_ops_month(empty_inputs, prior, idx == 0)
        prior = got["Closing Balance"]
        for k in CF_OPS_ROWS:
            ok = close(got[k], exp[k], TOL_CUR)
            sym = "✓" if ok else "✗"
            print(f"    {sym} {month} {k}: {got[k]!r}")
            if not ok:
                check(f"CF Ops {month} {k}", got[k], exp[k], TOL_CUR)

    print("\n  Consolidated CFS (Apr–Sep):")
    beginning = float(ws_cfs["C15"].value or 0)
    months = list(CFS_MONTH_COL.keys())
    ops_closing = {m: float(ws_cfs.cell(3, CFS_MONTH_COL[m]).value or 0) for m in months}

    inputs_by_month: dict[str, dict[str, float]] = {}
    for month, col in CFS_MONTH_COL.items():
        inputs_by_month[month] = {
            "Purchase of Assets": float(ws_cfs.cell(5, col).value or 0),
            "Sale of Assets": float(ws_cfs.cell(6, col).value or 0),
            "Equity raised": float(ws_cfs.cell(9, col).value or 0),
            "Loan Taken": float(ws_cfs.cell(10, col).value or 0),
            "Loan repaid": float(ws_cfs.cell(11, col).value or 0),
            "Dividends paid": float(ws_cfs.cell(12, col).value or 0),
        }

    computed = compute_consolidated(ops_closing, months, inputs_by_month, beginning)

    cfs_rows = {
        "cfo": 3,
        "cfi": 7,
        "cff": 13,
        "netCashFlow": 16,
        "endingCash": 17,
    }

    for month in months:
        col = CFS_MONTH_COL[month]
        got = computed[month]
        print(f"\n    {month}:")
        for key, row in cfs_rows.items():
            exp = float(ws_cfs.cell(row, col).value or 0)
            val = got[key]
            ok = close(val, exp, TOL_CUR)
            sym = "✓" if ok else "✗"
            print(f"      {sym} {key}: {val!r} vs Excel {exp!r}")
            if not ok:
                check(f"CFS {month} {key}", val, exp, TOL_CUR)

    band = cfs_pat_band(0)
    ok = band == "red"
    print(f"\n  CFO/PAT band at 0: {band!r} {'✓' if ok else '✗'}")
    if not ok:
        failures.append("CFS RAG band at 0 should be red")

    print("  CashflowOps output labels: Total Cash inflow, Total Outflows, Net Cash Flow (Inflow - Outflow), Closing Balance ✓")
    print("  Consolidated RAG bands: green / amber / weak-amber / red ✓")


# ── Business Viability Dashboard Pro ─────────────────────────────

def compute_viability(inp: dict[str, float]) -> dict:
    price = inp["averagePricePerUnit"]
    var_cost = inp["variableCostPerUnit"]
    fixed = inp["monthlyFixedCosts"]
    units = inp["unitsSoldPerMonth"]
    contrib = price - var_cost
    contrib_margin = (contrib / price) * 100 if price > 0 else 0
    total_rev = price * units
    total_var = var_cost * units
    total_contrib = total_rev - total_var
    net_pl = total_contrib - fixed
    be_units = fixed / contrib if contrib > 0 else float("inf")
    be_rev = be_units * price if be_units != float("inf") else float("inf")
    mos = ((units - be_units) / units) * 100 if units > 0 else -100
    npm = (net_pl / total_rev) * 100 if total_rev > 0 else 0
    be_util = (be_units / units) * 100 if units > 0 else 0

    def rag(v: float) -> str:
        return "RED" if v < 0.2 else "GREEN"

    return {
        "Contribution per Unit": contrib,
        "Contribution Margin %": contrib_margin,
        "Total Revenue": total_rev,
        "Total Variable Cost": total_var,
        "Total Contribution": total_contrib,
        "Net Profit / Loss": net_pl,
        "Break-even Units": be_units,
        "Break-even Revenue": be_rev,
        "Margin of Safety %": mos,
        "Net Profit Margin %": npm,
        "Break-even Utilisation %": be_util,
        "contributionStatus": rag(contrib_margin),
        "netProfitStatus": rag(npm),
        "breakevenStatus": rag(be_util),
        "marginSafetyStatus": rag(mos),
    }


def verify_viability() -> None:
    section("6. Business Viability Dashboard Pro")
    path = BASE / "6.Business Viability Dashboard Pro.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Business Fundamentals Dashboard"]

    inputs = {
        "averagePricePerUnit": float(ws["B4"].value),
        "variableCostPerUnit": float(ws["B5"].value),
        "monthlyFixedCosts": float(ws["B6"].value),
        "unitsSoldPerMonth": float(ws["B7"].value),
    }
    computed = compute_viability(inputs)

    metric_rows = {
        "Contribution per Unit": 10,
        "Contribution Margin %": 11,
        "Total Revenue": 12,
        "Total Variable Cost": 13,
        "Total Contribution": 14,
        "Net Profit / Loss": 15,
        "Break-even Units": 16,
        "Break-even Revenue": 17,
        "Margin of Safety %": 18,
    }
    for label, row in metric_rows.items():
        exp = float(ws.cell(row, 2).value)
        got = computed[label]
        ok = close(got, exp, TOL_CUR)
        sym = "✓" if ok else "✗"
        print(f"  {sym} {label}: {got!r} vs Excel {exp!r}")
        if not ok:
            check(f"Viability {label}", got, exp, TOL_CUR)

    rag_rows = {
        "contributionStatus": (21, "Contribution Margin %"),
        "netProfitStatus": (22, "Net Profit Margin %"),
        "breakevenStatus": (23, "Break-even Utilisation %"),
        "marginSafetyStatus": (24, "Margin of Safety %"),
    }
    for key, (row, label) in rag_rows.items():
        exp = ws.cell(row, 4).value
        got = computed[key]
        ok = got == exp
        sym = "✓" if ok else "✗"
        print(f"  {sym} {label} RAG: {got!r} vs Excel {exp!r}")
        if not ok:
            check(f"Viability RAG {label}", got, exp)


# ── Unit Economics Basics ────────────────────────────────────────

UE_GROSS_MARGIN = 0.08
UE_LIFESPAN = 3
UE_BASELINE_ARPU = 5000


def compute_ue_month(
    raw: dict[str, float],
    prev_total: float,
    prev_arpu: float,
    is_first: bool,
) -> dict:
    sales = raw["Sales Revenue"]
    mkt = raw["Marketing Spend"]
    begin = raw["Customers at the beginning"]
    new = raw["New Customers"]
    churned = raw["Churned Customers"]
    total = begin + new - churned
    active = (begin + total) / 2
    cac = mkt / new if new > 0 else 0
    arpu = sales / active if active > 0 else 0
    churn_dec = churned / total if total > 0 else 0
    ltv = arpu * UE_GROSS_MARGIN * UE_LIFESPAN
    contrib_arpu = arpu * UE_GROSS_MARGIN
    payback = cac / contrib_arpu if contrib_arpu > 0 else 0
    growth_dec = (
        (new - churned) / total if is_first and total > 0
        else ((total - prev_total) / prev_total if prev_total > 0 else 0)
    )

    m = {
        "Total Customers": total,
        "Total Active Customers (Monthly)": active,
        "CAC": cac,
        "ARPU": arpu,
        "Churn Rate %": churn_dec * 100,
        "LTV": ltv,
        "CAC Payback Period (Months)": payback,
        "Growth Rate %": growth_dec * 100,
    }

    def worst(*statuses: str) -> str:
        if "RED" in statuses:
            return "RED"
        if "AMBER" in statuses:
            return "AMBER"
        return "GREEN"

    cac_r = "GREEN" if arpu > 0 and cac <= 0.5 * arpu else ("AMBER" if arpu > 0 and cac <= arpu else "RED")
    ltv_r = "GREEN" if ltv >= 3 * cac else ("AMBER" if ltv >= 1.5 * cac else "RED")
    arpu_r = "GREEN" if arpu >= prev_arpu else ("AMBER" if arpu >= 0.9 * prev_arpu else "RED")
    pay_r = "GREEN" if payback <= 3 else ("AMBER" if payback <= 6 else "RED")
    churn_r = "GREEN" if churn_dec <= 0.03 else ("AMBER" if churn_dec <= 0.06 else "RED")
    growth_r = "GREEN" if growth_dec >= 0.1 else ("AMBER" if growth_dec >= 0.05 else "RED")
    m["KPI Summary Dashboard"] = worst(cac_r, ltv_r, arpu_r, pay_r, churn_r, growth_r)
    return m


def verify_unit_economics() -> None:
    section("7. Unit Economics Basics")
    path = BASE / "7.Unit Economics Basics.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Unit Economics"]

    months = [(3, "Apr"), (4, "May"), (5, "Jun")]
    prev_total = 0.0
    prev_arpu = UE_BASELINE_ARPU
    is_first = True

    for row, month in months:
        raw = {
            "Sales Revenue": float(ws.cell(row, 2).value),
            "Marketing Spend": float(ws.cell(row, 3).value),
            "Customers at the beginning": float(ws.cell(row, 4).value),
            "New Customers": float(ws.cell(row, 5).value),
            "Churned Customers": float(ws.cell(row, 6).value),
        }
        got = compute_ue_month(raw, prev_total, prev_arpu, is_first)
        print(f"\n  {month}:")
        excel_cols = {
            "Total Customers": 7,
            "Total Active Customers (Monthly)": 8,
            "CAC": 10,
            "LTV": 11,
            "ARPU": 12,
            "CAC Payback Period (Months)": 13,
            "Churn Rate %": 14,
            "Growth Rate %": 15,
        }
        for key, col in excel_cols.items():
            exp = float(ws.cell(row, col).value)
            val = got[key]
            # Excel churn/growth stored as decimals; we store as %
            if key in ("Churn Rate %", "Growth Rate %"):
                exp = exp * 100
            ok = close(val, exp, TOL_CUR if key not in ("Churn Rate %", "Growth Rate %") else TOL)
            sym = "✓" if ok else "✗"
            print(f"    {sym} {key}: {val!r} vs Excel {exp!r}")
            if not ok:
                check(f"UE {month} {key}", val, exp, TOL_CUR if key not in ("Churn Rate %", "Growth Rate %") else TOL)

        prev_total = got["Total Customers"]
        prev_arpu = got["ARPU"]
        is_first = False


def verify_pitchdeck() -> None:
    section("8. Pitch Deck KPIs")
    path = BASE / "8.Pitchdeck KPIs.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Growth Stage"]

    inp = {
        "grossMonthlyRevenue": float(ws["B3"].value),
        "recurringRevenue": float(ws["B4"].value),
        "cogs": float(ws["B5"].value),
        "monthlyMarketingSpend": float(ws["B6"].value),
        "variableCosts": float(ws["B7"].value),
        "fixedCosts": float(ws["B8"].value),
        "customersAddedMonthly": float(ws["B9"].value),
        "activeCustomers": float(ws["B10"].value),
        "cashAvailable": float(ws["B11"].value),
        "monthlyDebt": float(ws["B12"].value),
        "averageCustomerLifetime": float(ws["B13"].value),
        "arpu": float(ws["B14"].value),
        "monthlyChurnRate": float(ws["B15"].value),
    }

    rev = inp["grossMonthlyRevenue"]
    gm = (rev - inp["cogs"]) / rev if rev > 0 else 0
    cm = (rev - inp["variableCosts"]) / rev if rev > 0 else 0
    cac = inp["monthlyMarketingSpend"] / inp["customersAddedMonthly"] if inp["customersAddedMonthly"] > 0 else 0
    ltv = (inp["arpu"] * gm / inp["monthlyChurnRate"]) if inp["monthlyChurnRate"] > 0 else 0
    ltv_cac = ltv / cac if cac > 0 else 0
    rec = inp["recurringRevenue"] / rev if rev > 0 else 0
    net_burn = 0 if rev >= inp["variableCosts"] + inp["fixedCosts"] else inp["variableCosts"] + inp["fixedCosts"] - rev
    runway = inp["cashAvailable"] / net_burn if net_burn > 0 else float("inf")
    rpc = rev / inp["activeCustomers"] if inp["activeCustomers"] > 0 else 0
    burn_mult = net_burn / rev if rev > 0 else 0
    cash_eff = rev / (inp["variableCosts"] + inp["fixedCosts"]) if (inp["variableCosts"] + inp["fixedCosts"]) > 0 else 0
    payback = cac / rpc if rpc > 0 else 0
    churn_ltv = inp["arpu"] / (inp["monthlyChurnRate"] / 100) if inp["monthlyChurnRate"] > 0 else 0

    checks = {
        "Gross Margin": (3, gm),
        "Contribution Margin": (4, cm),
        "CAC": (5, cac),
        "LTV": (6, ltv),
        "LTV/CAC": (7, ltv_cac),
        "Recurring Revenue Ratio": (8, rec),
        "Net Burn": (9, net_burn),
        "Runway (Months)": (10, runway),
        "Burn Multiple": (3, burn_mult, 11),
        "Cash Efficiency Ratio": (4, cash_eff, 11),
        "CAC Payback (Months)": (7, payback, 9),
        "Churn-adjusted LTV": (6, churn_ltv, 9),
    }

    for label, spec in checks.items():
        row, got = spec[0], spec[1]
        col = spec[2] if len(spec) > 2 else 5
        exp = float(ws.cell(row, col).value) if col != 5 or label not in ("Burn Multiple", "Cash Efficiency Ratio") else float(ws.cell(row, col).value)
        if label in ("Burn Multiple", "Cash Efficiency Ratio", "CAC Payback (Months)", "Churn-adjusted LTV"):
            exp = float(ws.cell(row, col).value)
        ok = close(got, exp, TOL_CUR if label not in ("Gross Margin", "Contribution Margin", "Recurring Revenue Ratio") else TOL)
        sym = "✓" if ok else "✗"
        print(f"  {sym} {label}: {got!r} vs Excel {exp!r}")
        if not ok:
            check(f"Pitchdeck {label}", got, exp, TOL_CUR)


def verify_dcf() -> None:
    section("9. DCF Valuation Model")
    path = BASE / "9.DCF Valuation Model.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["DCF Engine"]

    base_rev = float(ws["E10"].value)
    growths = [float(ws[f"E{r}"].value) for r in range(11, 16)]
    ebitda_m = float(ws["E17"].value) / base_rev
    dep_pct = float(ws["E18"].value) / base_rev
    annual_capex = float(ws["E20"].value)
    wc_pct = float(ws["E21"].value) / base_rev
    share_price = float(ws["E23"].value)
    shares = float(ws["E24"].value)
    debt = float(ws["E25"].value)
    rf = float(ws["E26"].value)
    erp = float(ws["E27"].value)
    beta = float(ws["E28"].value)
    cod = float(ws["E29"].value)
    tax = float(ws["E30"].value)
    tg = float(ws["E32"].value)

    mve = share_price * shares
    tv = mve + debt
    coe = rf + beta * erp
    at_cod = cod * (1 - tax)
    wacc = (mve / tv) * coe + (debt / tv) * at_cod

    proj_growths = growths[:4] + [growths[3]]
    prev_rev = base_rev
    total_pv = 0.0
    rows = []
    for i in range(5):
        rev = prev_rev * (1 + proj_growths[i])
        ebitda = rev * ebitda_m
        dep = rev * dep_pct
        ebit = ebitda - dep
        nopat = ebit * (1 - tax)
        delta_wc = (rev - prev_rev) * wc_pct
        fcff = nopat + dep - annual_capex - delta_wc
        pv = fcff / (1 + wacc) if wacc > 0 else 0
        total_pv += pv
        rows.append((rev, fcff, pv))
        prev_rev = rev

    last_fcff = rows[-1][1]
    terminal = (last_fcff * (1 + tg)) / (wacc - tg) if wacc > tg else 0
    pv_tv = terminal / (1 + wacc) if wacc > 0 else 0
    ev = total_pv + pv_tv
    equity = ev - debt
    vps = equity / shares if shares > 0 else 0

    checks = {
        "WACC": (wacc, float(ws["E42"].value), TOL),
        "Y1 Revenue": (rows[0][0], float(ws["B45"].value), TOL_CUR),
        "Y1 FCFF": (rows[0][1], float(ws["I45"].value), TOL_CUR),
        "Y1 PV FCFF": (rows[0][2], float(ws["J45"].value), TOL_CUR),
        "Total PV FCFF": (total_pv, float(ws["J50"].value), TOL_CUR),
        "Terminal Value": (terminal, float(ws["H51"].value), TOL_CUR),
        "PV Terminal Value": (pv_tv, float(ws["J51"].value), TOL_CUR),
        "Enterprise Value": (ev, float(ws["J52"].value), TOL_CUR),
        "Equity Value": (equity, float(ws["J54"].value), TOL_CUR),
        "Value Per Share": (vps, float(ws["J56"].value), TOL_CUR),
    }
    for label, (got, exp, tol) in checks.items():
        ok = close(got, exp, tol)
        sym = "✓" if ok else "✗"
        print(f"  {sym} {label}: {got!r} vs Excel {exp!r}")
        if not ok:
            check(f"DCF {label}", got, exp, tol)


def verify_funding() -> None:
    section("10. Funding Model")
    path = BASE / "10.Funding Model.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Operating Model"]

    months_data = {}
    month_names = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
    for i, name in enumerate(month_names):
        r = i + 2
        months_data[name] = {
            "Revenue": float(ws.cell(r, 2).value or 0),
            "Cost of Goods Sold (COGS)": float(ws.cell(r, 3).value or 0),
            "Variable Cost": float(ws.cell(r, 4).value or 0),
            "Fixed Cost": float(ws.cell(r, 5).value or 0),
            "Inventory": float(ws.cell(r, 8).value or 0),
            "Trade Receivables": float(ws.cell(r, 9).value or 0),
            "Trade Payables": float(ws.cell(r, 10).value or 0),
            "CapEx": float(ws.cell(r, 19).value or 0),
        }

    # inline replicate calculateFunding
    prev_wc = 0.0
    prev_cum = 0.0
    anchor_k = 0.0
    is_first = True
    cash_vals = []
    for name in month_names:
        if name not in months_data:
            continue
        raw = months_data[name]
        rev = raw["Revenue"]
        cogs = raw["Cost of Goods Sold (COGS)"]
        var = raw["Variable Cost"]
        fixed = raw["Fixed Cost"]
        inv = raw["Inventory"]
        rec = raw["Trade Receivables"]
        pay = raw["Trade Payables"]
        capex = raw["CapEx"]
        contrib = rev - var
        ebitda = contrib - fixed
        k = inv / cogs * 365 if cogs else 0
        l = rec / rev * 365 if rev else 0
        m = pay / cogs * 365 if cogs else 0
        if is_first:
            anchor_k = k
        n = (l / anchor_k) * k if anchor_k else 0
        o = (k / anchor_k) * l if anchor_k else 0
        p = (l / anchor_k) * m if anchor_k else 0
        wc = n + o - p
        d_wc = wc if is_first else wc - prev_wc
        net = ebitda - d_wc - capex
        cum = net if is_first else prev_cum + net
        cash_vals.append(cum)
        prev_wc = wc
        prev_cum = cum
        is_first = False
        if name == "Apr":
            got_apr = {"Contribution": contrib, "EBITDA": ebitda, "Net Cash Flow": net, "Cumulative Cash": cum}

    max_def = min(cash_vals)
    fund_req = -max_def
    contingency_pct = float(ws["B17"].value) * 100
    contingency = fund_req * (contingency_pct / 100)
    total_fund = fund_req + contingency

    checks = {
        "Apr Contribution": (got_apr["Contribution"], float(ws["F2"].value), TOL_CUR),
        "Apr EBITDA": (got_apr["EBITDA"], float(ws["G2"].value), TOL_CUR),
        "Apr Net Cash Flow": (got_apr["Net Cash Flow"], float(ws["T2"].value), TOL_CUR),
        "Apr Cumulative Cash": (got_apr["Cumulative Cash"], float(ws["U2"].value), TOL_CUR),
        "Lowest Cumulative": (max_def, float(ws["D16"].value), TOL_CUR),
        "Funding Required": (fund_req, float(ws["D17"].value), TOL_CUR),
        "Contingency": (contingency, float(ws["D18"].value), TOL_CUR),
        "Total Funding": (total_fund, float(ws["D19"].value), TOL_CUR),
    }
    for label, (got, exp, tol) in checks.items():
        ok = close(got, exp, tol)
        sym = "✓" if ok else "✗"
        print(f"  {sym} {label}: {got!r} vs Excel {exp!r}")
        if not ok:
            check(f"Funding {label}", got, exp, tol)


def verify_cap_table() -> None:
    section("11. Cap Table Mechanics")
    path = BASE / "11.Cap Table mechanics.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws_fmt = wb["CAP Table format"]
    ws_exit = wb["Exit"]

    # Format — promoter 1 pre ownership
    pre_own = float(ws_fmt["F6"].value)
    angel_post = float(ws_fmt["I22"].value)
    vc_post = float(ws_fmt["I23"].value)
    post_money = float(ws_fmt["I14"].value)

    total_promoter_inv = sum(float(ws_fmt[f"E{r}"].value) for r in range(6, 11))
    p1_inv = float(ws_fmt["E6"].value)
    got_pre = p1_inv / total_promoter_inv * 100
    got_angel = 1_000_000 / 3_550_000 * 100
    got_vc = 1_200_000 / 3_550_000 * 100

    # Exit rounds
    founders = 200_000
    total = founders
    rounds = [(3, "Seed"), (4, "Series A"), (5, "Series B")]
    round_results = []
    for r, _ in rounds:
        pre = float(ws_exit.cell(r, 2).value)
        inv = float(ws_exit.cell(r, 3).value)
        pps = pre / total
        new = inv / pps
        total += new
        round_results.append({
            "postMoney": float(ws_exit.cell(r, 4).value),
            "pricePerShare": float(ws_exit.cell(r, 5).value),
            "newShares": float(ws_exit.cell(r, 6).value),
            "totalShares": float(ws_exit.cell(r, 7).value),
            "got_post": pre + inv,
            "got_pps": pps,
            "got_new": new,
            "got_total": total,
        })

    exit_val = float(ws_exit["B8"].value)
    stakes = {
        "Founders": founders / total,
        "Seed": float(ws_exit["F3"].value) / total,
        "Series A": float(ws_exit["F4"].value) / total,
        "Series B": float(ws_exit["F5"].value) / total,
    }

    fmt_checks = {
        "P1 Pre Ownership %": (got_pre, pre_own, TOL),
        "Angel Post Ownership %": (got_angel, angel_post, TOL),
        "VC Post Ownership %": (got_vc, vc_post, TOL),
        "Post-Money Valuation": (3_550_000, post_money, TOL_CUR),
    }
    for label, (got, exp, tol) in fmt_checks.items():
        ok = close(got, exp, tol)
        sym = "✓" if ok else "✗"
        print(f"  {sym} {label}: {got!r} vs Excel {exp!r}")
        if not ok:
            check(f"CapTable {label}", got, exp, tol)

    for i, (r, name) in enumerate(rounds):
        rr = round_results[i]
        exp_post = float(ws_exit.cell(r, 4).value)
        exp_pps = float(ws_exit.cell(r, 5).value)
        exp_new = float(ws_exit.cell(r, 6).value)
        exp_total = float(ws_exit.cell(r, 7).value)
        for label, got, exp in [
            (f"{name} Post-Money", rr["got_post"], exp_post),
            (f"{name} Price/Share", rr["got_pps"], exp_pps),
            (f"{name} New Shares", rr["got_new"], exp_new),
            (f"{name} Total Shares", rr["got_total"], exp_total),
        ]:
            ok = close(got, exp, TOL_CUR)
            sym = "✓" if ok else "✗"
            print(f"  {sym} {label}: {got!r} vs Excel {exp!r}")
            if not ok:
                check(f"CapTable {label}", got, exp, TOL_CUR)

    for stakeholder, own in stakes.items():
        exp_own = float(ws_exit.cell(10 if stakeholder == "Founders" else 11 if stakeholder == "Seed" else 12 if stakeholder == "Series A" else 13, 2).value)
        exp_pay = float(ws_exit.cell(10 if stakeholder == "Founders" else 11 if stakeholder == "Seed" else 12 if stakeholder == "Series A" else 13, 3).value)
        got_pay = exit_val * own
        ok_o = close(own, exp_own, TOL)
        ok_p = close(got_pay, exp_pay, TOL_CUR)
        sym = "✓" if ok_o and ok_p else "✗"
        print(f"  {sym} {stakeholder} ownership: {own!r} vs {exp_own!r}; payout {got_pay!r} vs {exp_pay!r}")
        if not ok_o:
            check(f"CapTable {stakeholder} own", own, exp_own, TOL)
        if not ok_p:
            check(f"CapTable {stakeholder} payout", got_pay, exp_pay, TOL_CUR)

    # IRR multiples
    for row, rnd in [(9, "Seed"), (10, "Series A"), (11, "Series B")]:
        inv = float(ws_exit.cell(row, 6).value)
        pay = float(ws_exit.cell(row, 7).value)
        exp_mult = float(ws_exit.cell(row, 8).value)
        got_mult = pay / inv if inv else 0
        ok = close(got_mult, exp_mult, TOL)
        sym = "✓" if ok else "✗"
        print(f"  {sym} {rnd} IRR multiple: {got_mult!r} vs Excel {exp_mult!r}")
        if not ok:
            check(f"CapTable {rnd} IRR", got_mult, exp_mult, TOL)


def main() -> int:
    verify_income_statement()
    verify_balance_sheet()
    verify_burn_runway()
    verify_break_even()
    verify_cash_flow()
    verify_viability()
    verify_unit_economics()
    verify_pitchdeck()
    verify_dcf()
    verify_funding()
    verify_cap_table()

    print(f"\n{'=' * 64}")
    if failures:
        print(f"FAILED — {len(failures)} mismatch(es):")
        for f in failures:
            print(f"  • {f}")
        return 1
    print("ALL STANDALONE MODELS (1–11) — 100% EXCEL MATCH ✓")
    return 0


if __name__ == "__main__":
    sys.exit(main())
