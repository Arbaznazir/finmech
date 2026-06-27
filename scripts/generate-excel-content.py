#!/usr/bin/env python3
"""Generate frontend/src/lib/excel-exact-content.generated.ts from FINMECH-UPGRADED workbooks."""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "FINMECH-UPGRADED/2.Stand alone models"
OUT = ROOT / "frontend/src/lib/excel-exact-content.generated.ts"


def cell(ws, r: int, c: int) -> str | None:
    v = ws.cell(r, c).value
    if v is None:
        return None
    s = str(v).strip()
    return s if s and not s.startswith("=") else None


def esc(s: str | None) -> str:
    if s is None:
        return ""
    return (
        str(s)
        .replace("\\", "\\\\")
        .replace("`", "\\`")
        .replace("${", "\\${")
    )


def extract_if_string(s: str | None) -> str | None:
    if not s:
        return None
    m = re.search(r'"([^"]{20,})"', s)
    return m.group(1) if m else s


def ts_string(s: str | None) -> str:
    if s is None:
        return '""'
    return f"`{esc(s)}`"


def ts_key(k: str) -> str:
    return json.dumps(k)


def ts_object_pairs(pairs: dict[str, str], indent: str = "  ") -> list[str]:
    return [f"{indent}{ts_key(k)}: {ts_string(v)}," for k, v in pairs.items()]


def main() -> None:
    lines: list[str] = [
        "// AUTO-GENERATED from FINMECH-UPGRADED Excel files. Do not edit by hand.",
        "// Run: python3 scripts/generate-excel-content.py",
        "",
    ]

    # Income Statement
    ws = openpyxl.load_workbook(BASE / "1.Income Statement.xlsx", data_only=True)["Income Statement"]
    rag_by_color: dict[str, str] = {}
    for c in range(1, 25):
        v = cell(ws, 43, c)
        if v and "-" in v:
            color = v.split("-", 1)[0].strip().upper()
            rag_by_color.setdefault(color, v)
    mentoring: list[str] = []
    seen: set[str] = set()
    for r in range(98, ws.max_row + 1):
        for c in range(1, 25):
            v = cell(ws, r, c)
            if v and len(v) > 25 and "MENTORING" not in v.upper() and v not in seen:
                seen.add(v)
                mentoring.append(v)
    lines += [
        "export const IS_EXACT = {",
        "  ragByColor: {",
        *[f"    {k}: {ts_string(v)}," for k, v in rag_by_color.items()],
        "  } as const,",
        "  mentoringGuidance: [",
        *[f"    {ts_string(m)}," for m in mentoring],
        "  ],",
        "};",
        "",
    ]

    # Balance Sheet
    ws = openpyxl.load_workbook(BASE / "2.Balance Sheet.xlsx", data_only=True)["Balance Sheet"]
    lines += [
        "export const BS_EXACT = {",
        f"  tallyBalanced: {ts_string(cell(ws, 39, 4))},",
        f"  tallyUnbalanced: {ts_string(cell(ws, 39, 3))},",
        f"  scenarioGreen: {ts_string(cell(ws, 46, 4))},",
        f"  analyticsGreen: {ts_string(cell(ws, 50, 4))},",
        f"  analyticsRed: {ts_string(cell(ws, 52, 3))},",
        f"  structuralRed: {ts_string(cell(ws, 48, 3))},",
        "  action: {",
        *[
            f"    {color}: {ts_string(extract_if_string(cell(ws, row, 4)))},"
            for color, row in [("GREEN", 53), ("AMBER", 54), ("RED", 55)]
        ],
        "  },",
        "};",
        "",
    ]

    # Burn & Runway
    ws = openpyxl.load_workbook(BASE / "3.Burn and Runway Monitor.xlsx", data_only=True)["Budget"]
    h = {c: cell(ws, 3, c) for c in range(1, 25)}
    field_what = {h[c]: cell(ws, 16, c) for c in range(2, 18) if h.get(c) and cell(ws, 16, c)}
    field_why = {h[c]: cell(ws, 17, c) for c in range(2, 18) if h.get(c) and cell(ws, 17, c)}
    lines += [
        "export const BURN_EXACT = {",
        "  fieldWhat: {",
        *[f"    {ts_key(k)}: {ts_string(v)}," for k, v in field_what.items()],
        "  },",
        "  fieldWhy: {",
        *[f"    {ts_key(k)}: {ts_string(v)}," for k, v in field_why.items()],
        "  },",
        "  classification: {",
        "    GREEN: {",
        f"      overall: {ts_string(cell(ws, 4, 19))},",
        "      guidance: [",
        f"        {ts_string(cell(ws, 4, 22))},",
        "      ],",
        "    },",
        "    AMBER: {",
        f"      overall: {ts_string(cell(ws, 27, 18))},",
        "      guidance: [",
        f"        {ts_string(cell(ws, 22, 20))},",
        "      ],",
        "    },",
        "    RED: {",
        f"      overall: {ts_string(cell(ws, 5, 21))},",
        "      guidance: [",
        f"        {ts_string(cell(ws, 5, 24))},",
        "      ],",
        "    },",
        "  },",
        "};",
        "",
    ]

    # Break-even
    ws = openpyxl.load_workbook(BASE / "5.Break-even Model.xlsx", data_only=True)["BreakEven Table"]
    be_rows = {
        "pricePerUnit": 3,
        "variableCostPerUnit": 4,
        "fixedCostMonthly": 5,
        "contributionPerUnit": 7,
        "breakEvenUnits": 8,
        "breakEvenRevenue": 9,
    }
    lines += ["export const BREAK_EVEN_EXACT = {"]
    for key, r in be_rows.items():
        lines.append(f"  {key}: {{ question: {ts_string(cell(ws, r, 3))}, what: {ts_string(cell(ws, r, 4))} }},")
    lines += ["};", ""]

    # Viability
    ws = openpyxl.load_workbook(BASE / "6.Business Viability Dashboard Pro.xlsx", data_only=True)[
        "Business Fundamentals Dashboard"
    ]
    lines += ["export const VIABILITY_EXACT: Record<string, Record<string, { meaning: string; mentoring: string }>> = {"]
    for r in range(21, 30):
        metric, status, meaning, mentoring = cell(ws, r, 1), cell(ws, r, 4), cell(ws, r, 5), cell(ws, r, 6)
        if metric and status and meaning:
            lines += [
                f"  {ts_key(metric)}: {{",
                f"    {status}: {{ meaning: {ts_string(meaning)}, mentoring: {ts_string(mentoring or '')} }},",
                "  },",
            ]
    lines += ["};", ""]

    # Unit Economics
    ws = openpyxl.load_workbook(BASE / "7.Unit Economics Basics.xlsx", data_only=True)["Unit Economics"]
    h2 = {c: cell(ws, 2, c) for c in range(1, 16)}
    what = {h2[c]: cell(ws, 19, c) for c in range(2, 16) if h2.get(c) and cell(ws, 19, c) not in (None, "Note")}
    why = {h2[c]: cell(ws, 20, c) for c in range(2, 16) if h2.get(c) and cell(ws, 20, c) not in (None, "Why it matters")}
    rag: dict[str, dict[str, str]] = {}
    for r in range(21, 45):
        for c in range(10, 16):
            v, hn = cell(ws, r, c), h2.get(c)
            if v and hn and ":" in v:
                color, msg = v.split(":", 1)
                rag.setdefault(hn, {})[color.strip().upper()] = msg.strip()
    lines += [
        "export const UE_EXACT = {",
        "  fieldWhat: {",
        *[f"    {ts_key(k)}: {ts_string(v)}," for k, v in what.items()],
        "  },",
        "  fieldWhy: {",
        *[f"    {ts_key(k)}: {ts_string(v)}," for k, v in why.items()],
        "  },",
        "  ragCommentary: {",
    ]
    for metric, colors in rag.items():
        lines.append(f"    {ts_key(metric)}: {{")
        for color, msg in colors.items():
            lines.append(f"      {color}: {ts_string(msg)},")
        lines.append("    },")
    lines += ["  },", "};", ""]

    # Pitch Deck
    ws = openpyxl.load_workbook(BASE / "8.Pitchdeck KPIs.xlsx", data_only=True)["Growth Stage"]
    smart_lines: list[str] = []
    metric_commentary: dict[str, str] = {}
    for r in range(13, 21):
        for metric_c, comment_c in [(4, 5), (6, 7), (8, 9), (10, 11)]:
            m, c = cell(ws, r, metric_c), cell(ws, r, comment_c)
            if m and c:
                line = f"{m}: {c}"
                smart_lines.append(line)
                if metric_c == 4 and ":" in c:
                    metric_commentary[m] = c
    ws_def = openpyxl.load_workbook(BASE / "8.Pitchdeck KPIs.xlsx", data_only=True)["Definitions"]
    definitions: dict[str, str] = {}
    for r in range(1, 40):
        a, b = cell(ws_def, r, 1), cell(ws_def, r, 2)
        if b and len(b) > 15 and not (a and a.isdigit()):
            key = (a or b).split("(")[0].strip().lower().replace(" ", "_").replace("/", "_")
            key = re.sub(r"_+", "_", key).strip("_")
            if key and not key.isdigit():
                definitions[key] = b
    lines += [
        "export const PITCHDECK_EXACT = {",
        "  metricCommentary: {",
        *[f"    {ts_key(k)}: {ts_string(v)}," for k, v in metric_commentary.items()],
        "  },",
        "  smartReportLines: [",
        *[f"    {ts_string(line)}," for line in smart_lines],
        "  ],",
        "  definitions: {",
        *[f"    {k}: {ts_string(v)}," for k, v in definitions.items() if k.isidentifier()],
        "  },",
        "};",
        "",
    ]

    # Cash Flow (Consolidated CFS interpretation table)
    ws = openpyxl.load_workbook(BASE / "4.Cash Flow Statement.xlsx", data_only=True)["Consolidated CFS"]
    cfo_pat_rows: list[tuple[str, str, str]] = []
    cfo_ebitda_rows: list[tuple[str, str, str]] = []
    for r in range(26, 30):
        pat_range, pat_colour, pat_meaning = cell(ws, r, 2), cell(ws, r, 3), cell(ws, r, 4)
        if pat_range and pat_meaning:
            cfo_pat_rows.append((pat_range, pat_colour or "", pat_meaning))
        ebitda_range, ebitda_colour, ebitda_meaning = cell(ws, r, 8), cell(ws, r, 9), cell(ws, r, 10)
        if ebitda_range and ebitda_meaning:
            cfo_ebitda_rows.append((ebitda_range, ebitda_colour or "", ebitda_meaning))
    lines += [
        "export const CFS_EXACT = {",
        f"  strongOverall: {ts_string(cell(ws, 21, 7))},",
        "  cfoPatInterpretation: [",
        *[
            f"    {{ range: {ts_string(rng)}, colour: {ts_string(col)}, meaning: {ts_string(mean)} }},"
            for rng, col, mean in cfo_pat_rows
        ],
        "  ],",
        "  cfoEbitdaInterpretation: [",
        *[
            f"    {{ range: {ts_string(rng)}, colour: {ts_string(col)}, meaning: {ts_string(mean)} }},"
            for rng, col, mean in cfo_ebitda_rows
        ],
        "  ],",
        "};",
        "",
    ]

    # DCF
    ws = openpyxl.load_workbook(BASE / "9.DCF Valuation Model.xlsx", data_only=True)["DCF Engine"]
    lines += ["export const DCF_EXACT = {"]
    for prefix, row in [("mentoring", 62), ("interpretation", 61)]:
        for label, col in [("Base", 3), ("Upside", 4), ("Downside", 5)]:
            v = cell(ws, row, col)
            if v:
                lines.append(f"  {prefix}{label}: {ts_string(v)},")
    lines += ["};", ""]

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT} ({len(lines)} lines)")


if __name__ == "__main__":
    main()
